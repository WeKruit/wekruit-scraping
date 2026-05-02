#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import sys
import zipfile
from collections import OrderedDict
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

ROOT = Path(__file__).resolve().parents[1]
RESEARCHER_ROOT = ROOT / "researcher"
if str(RESEARCHER_ROOT) not in sys.path:
    sys.path.insert(0, str(RESEARCHER_ROOT))

from pipeline.sourcing_client import DEFAULT_BATCH_SIZE, dry_run_upload, upload_source_records
from pipeline.sourcing_records import content_hash, raw_storage_path, safe_id, utc_now

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - only reached when dependencies are missing.
    load_workbook = None


def first_non_empty(*values: Any) -> str:
    for value in values:
        if isinstance(value, list):
            joined = "; ".join(str(item).strip() for item in value if str(item).strip())
            if joined:
                return joined
        elif value not in (None, "", [], {}):
            text = str(value).strip()
            if text:
                return text
    return ""


def list_values(value: Any) -> list[str]:
    if isinstance(value, list):
        output: list[str] = []
        for item in value:
            output.extend(list_values(item))
        return output
    if value in (None, "", [], {}):
        return []
    text = str(value).strip()
    return [text] if text else []


def split_values(value: Any, *, separators: tuple[str, ...] = (";", "\n")) -> list[str]:
    output: list[str] = []
    for text in list_values(value):
        parts = [text]
        for separator in separators:
            next_parts: list[str] = []
            for part in parts:
                next_parts.extend(part.split(separator))
            parts = next_parts
        output.extend(part.strip() for part in parts if part.strip())
    return output


def unique_values(*values: Any) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for value in values:
        for text in list_values(value):
            if text not in seen:
                seen.add(text)
                output.append(text)
    return output


def unique_split_values(*values: Any, separators: tuple[str, ...] = (";", "\n")) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for value in values:
        for text in split_values(value, separators=separators):
            if text not in seen:
                seen.add(text)
                output.append(text)
    return output


def compact(payload: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in payload.items() if value not in (None, "", [], {})}


def safe_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def suggested_signals(*signals: str) -> list[str]:
    return unique_values(*signals)


def normalized_url_key(value: Any) -> str:
    text = first_non_empty(value)
    if not text:
        return ""
    try:
        parsed = urlparse(text)
        if parsed.scheme and parsed.netloc:
            return f"{parsed.netloc.lower()}{parsed.path.rstrip('/').lower()}"
    except ValueError:
        pass
    return text.strip().rstrip("/").lower()


def url_last_segment(value: Any) -> str:
    text = first_non_empty(value)
    if not text:
        return ""
    try:
        parsed = urlparse(text)
        path = parsed.path.rstrip("/")
        if path:
            return path.rsplit("/", 1)[-1]
    except ValueError:
        pass
    return text.rstrip("/").rsplit("/", 1)[-1]


def github_repo_ref(value: Any) -> str:
    text = first_non_empty(value)
    if not text:
        return ""
    try:
        parsed = urlparse(text)
        if "github.com" in parsed.netloc.lower():
            parts = [part for part in parsed.path.strip("/").split("/") if part]
            if len(parts) >= 2:
                return f"{parts[0]}/{parts[1]}"
    except ValueError:
        pass
    return text


def link_domain(value: Any) -> str:
    text = first_non_empty(value)
    if not text:
        return ""
    try:
        return urlparse(text).netloc.lower()
    except ValueError:
        return ""


def github_suggested_signals(candidate: dict[str, Any]) -> list[str]:
    profile = candidate.get("profile") if isinstance(candidate.get("profile"), dict) else {}
    repos = candidate.get("repos") if isinstance(candidate.get("repos"), list) else []
    total_commits = safe_int(candidate.get("total_commits"))
    merged_prs = safe_int(candidate.get("merged_prs"))
    followers = safe_int(profile.get("followers"))
    signals = ["open_source_contribution", "technical_project"]
    if total_commits >= 25 or merged_prs >= 5 or len(repos) >= 2:
        signals.append("high_activity")
    if merged_prs >= 10 or followers >= 100:
        signals.append("maintainer_or_influence_signal")
    return suggested_signals(*signals)


def devpost_project_signals(project: dict[str, Any]) -> list[str]:
    signals = ["technical_project", "hackathon_participation"]
    if project.get("winner") or project.get("prizes"):
        signals.append("award_or_recognition")
    if project.get("github_links") or project.get("github_repos"):
        signals.append("builder_signal")
    return suggested_signals(*signals)


def devpost_member_signals(rows: list[dict[str, Any]]) -> list[str]:
    signals = ["hackathon_participation", "technical_project", "founder_or_builder_signal"]
    if any(first_non_empty(row.get("winner"), row.get("prizes")) for row in rows):
        signals.append("award_or_recognition")
    if any(first_non_empty(row.get("member_github"), row.get("github_repos")) for row in rows):
        signals.append("open_source_contribution")
    return suggested_signals(*signals)


def is_devpost_flat_row(record: dict[str, Any]) -> bool:
    return any(
        key in record
        for key in (
            "member_username",
            "member_devpost",
            "member_github",
            "member_linkedin",
            "member_twitter",
            "member_website",
        )
    )


def devpost_member_key(record: dict[str, Any]) -> str:
    for key in ("member_devpost", "member_github", "member_linkedin", "member_username"):
        value = first_non_empty(record.get(key))
        if value:
            return f"{key}:{normalized_url_key(value)}"
    return f"row:{content_hash(record).split(':', 1)[1][:16]}"


def devpost_project_key(record: dict[str, Any]) -> str:
    project_url = first_non_empty(record.get("project_url"))
    if project_url:
        return f"url:{normalized_url_key(project_url)}"
    return f"name:{first_non_empty(record.get('hackathon'))}:{first_non_empty(record.get('project_name'))}".lower()


def devpost_project_context(row: dict[str, Any]) -> dict[str, Any]:
    github_repo_links = unique_split_values(row.get("github_repos"))
    demo_links = unique_split_values(row.get("demo_links"))
    all_links = unique_split_values(row.get("all_links"))
    return compact({
        "hackathon": row.get("hackathon"),
        "projectName": row.get("project_name"),
        "tagline": row.get("tagline"),
        "description": row.get("description"),
        "projectUrl": row.get("project_url"),
        "projectRef": url_last_segment(row.get("project_url")),
        "videoUrl": row.get("video_url"),
        "winner": row.get("winner"),
        "likes": row.get("likes"),
        "projectGithubRepos": github_repo_links,
        "projectGithubRepoRefs": unique_values([github_repo_ref(link) for link in github_repo_links]),
        "demoLinks": demo_links,
        "allLinks": all_links,
        "externalLinkDomains": unique_values(
            [link_domain(link) for link in [*demo_links, *all_links] if "github.com" not in link.lower()]
        ),
        "techStack": unique_split_values(row.get("tech_stack"), separators=(",", ";", "\n")),
        "prizes": unique_split_values(row.get("prizes")),
        "imageCount": row.get("image_count"),
        "inputFile": row.get("_input_file"),
    })


def devpost_flat_records(rows: Iterable[dict[str, Any]]) -> Iterable[dict[str, Any]]:
    grouped: "OrderedDict[str, list[dict[str, Any]]]" = OrderedDict()
    for row in rows:
        if not is_devpost_flat_row(row):
            continue
        if not any(
            first_non_empty(row.get(key))
            for key in (
                "member_name",
                "member_username",
                "member_devpost",
                "member_github",
                "member_linkedin",
                "member_twitter",
                "member_website",
            )
        ):
            continue
        grouped.setdefault(devpost_member_key(row), []).append(row)

    for member_rows in grouped.values():
        first = member_rows[0]
        member_name = first_non_empty(
            first.get("member_name"),
            first.get("name"),
            first.get("member_username"),
            url_last_segment(first.get("member_devpost")),
            url_last_segment(first.get("member_github")),
        )
        member_devpost = first_non_empty(*[row.get("member_devpost") for row in member_rows])
        member_github = first_non_empty(*[row.get("member_github") for row in member_rows])
        member_linkedin = first_non_empty(*[row.get("member_linkedin") for row in member_rows])
        member_twitter = first_non_empty(*[row.get("member_twitter") for row in member_rows])
        member_website = first_non_empty(*[row.get("member_website") for row in member_rows])

        projects_by_key: "OrderedDict[str, dict[str, Any]]" = OrderedDict()
        for row in member_rows:
            projects_by_key.setdefault(devpost_project_key(row), devpost_project_context(row))
        projects = list(projects_by_key.values())
        project_names = unique_values([project.get("projectName") for project in projects])
        hackathons = unique_values([project.get("hackathon") for project in projects])
        tech_stack = unique_values([tag for project in projects for tag in project.get("techStack", [])])
        prizes = unique_values([prize for project in projects for prize in project.get("prizes", [])])
        project_github_repo_refs = unique_values(
            [repo for project in projects for repo in project.get("projectGithubRepoRefs", [])]
        )

        yield {
            "entityType": "person",
            "sourceNativeId": first_non_empty(
                member_devpost,
                member_github,
                member_linkedin,
                first.get("member_username"),
                member_name,
            ),
            "sourceUrl": member_devpost,
            "displayName": member_name,
            "display": compact({
                "name": member_name,
                "homepage": member_website,
                "github": member_github,
                "linkedin": member_linkedin,
                "twitter": member_twitter,
            }),
            "rawSummary": compact({
                "devpost": member_devpost,
                "github": member_github,
                "linkedin": member_linkedin,
                "twitter": member_twitter,
                "homepage": member_website,
                "projectCount": len(projects),
                "projectNames": project_names[:10],
                "hackathons": hackathons[:10],
                "tech": tech_stack[:30],
                "prizes": prizes[:10],
                "winnerProjectCount": sum(1 for project in projects if project.get("winner")),
                "projectGithubRepoRefs": project_github_repo_refs[:20],
                "suggestedSignals": devpost_member_signals(member_rows),
            }),
            "raw": {
                "member": compact({
                    "name": member_name,
                    "username": first.get("member_username"),
                    "devpost": member_devpost,
                    "github": member_github,
                    "linkedin": member_linkedin,
                    "twitter": member_twitter,
                    "website": member_website,
                }),
                "projects": projects,
                "sourceRows": member_rows,
            },
        }


def research_suggested_signals(record: dict[str, Any]) -> list[str]:
    entity_type = str(record.get("entityType") or record.get("entity_type") or record.get("type") or "").lower()
    signals: list[str] = []
    if entity_type in {"paper", "research_work", "work", "works"} or record.get("doi") or record.get("paper_title"):
        signals.append("research_publication")
    if record.get("institution") or record.get("affiliation") or record.get("company"):
        signals.append("education_affiliation")
    if record.get("orcid") or record.get("openalex_author_id"):
        signals.append("academic_research_signal")
    return suggested_signals(*signals)


def _row_dict(headers: list[str], row: tuple[Any, ...], *, input_file: str = "") -> dict[str, Any]:
    record = {
        headers[index]: "" if index >= len(row) or row[index] is None else str(row[index]).strip()
        for index in range(len(headers))
        if headers[index]
    }
    if input_file:
        record["_input_file"] = input_file
    return record


def load_csv_text_records(text: str, *, input_file: str = "") -> list[dict[str, Any]]:
    records = [dict(row) for row in csv.DictReader(StringIO(text))]
    if input_file:
        for record in records:
            record["_input_file"] = input_file
    return records


def load_xlsx_records_from_bytes(payload: bytes, *, input_file: str = "") -> list[dict[str, Any]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read .xlsx files. Install dependencies from requirements.txt.")
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = ["" if header is None else str(header).strip() for header in next(rows)]
    except StopIteration:
        workbook.close()
        return []

    output: list[dict[str, Any]] = []
    for row in rows:
        record = _row_dict(headers, row, input_file=input_file)
        if any(value for key, value in record.items() if not key.startswith("_")):
            output.append(record)
    workbook.close()
    return output


def load_xlsx_records(path: Path) -> list[dict[str, Any]]:
    return load_xlsx_records_from_bytes(path.read_bytes(), input_file=str(path))


def load_zip_records(path: Path) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    with zipfile.ZipFile(path) as archive:
        for name in sorted(archive.namelist()):
            if name.endswith("/") or name.startswith("__MACOSX/"):
                continue
            suffix = Path(name).suffix.lower()
            if suffix == ".xlsx":
                output.extend(load_xlsx_records_from_bytes(archive.read(name), input_file=name))
            elif suffix == ".csv":
                output.extend(load_csv_text_records(archive.read(name).decode("utf-8-sig"), input_file=name))
    return output


def load_directory_records(path: Path) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for child in sorted(item for item in path.rglob("*") if item.is_file()):
        if child.name.startswith(".") or "__MACOSX" in child.parts:
            continue
        if child.suffix.lower() in {".csv", ".json", ".jsonl", ".xlsx", ".zip"}:
            output.extend(load_input_records(child))
    return output


def load_input_records(path: Path) -> list[dict[str, Any]]:
    if path.is_dir():
        return load_directory_records(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    if suffix == ".xlsx":
        return load_xlsx_records(path)
    if suffix == ".zip":
        return load_zip_records(path)
    if suffix == ".jsonl":
        return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
    if suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, list):
            return payload
        if isinstance(payload, dict):
            for key in ("records", "items", "projects", "candidates", "data"):
                value = payload.get(key)
                if isinstance(value, list):
                    return value
            return [payload]
    raise ValueError(f"Unsupported input file format: {path}")


def iter_domain_records(source: str, records: Iterable[dict[str, Any]]) -> Iterable[dict[str, Any]]:
    materialized = list(records)
    if source == "devpost" and any(is_devpost_flat_row(record) for record in materialized):
        yield from devpost_flat_records(materialized)
        return

    for record in materialized:
        if source == "devpost":
            yield from devpost_records(record)
        elif source == "github":
            yield github_record(record)
        else:
            yield generic_record(record)


def devpost_records(project: dict[str, Any]) -> Iterable[dict[str, Any]]:
    project_name = first_non_empty(project.get("name"), project.get("project_name"))
    project_url = first_non_empty(project.get("project_url"), project.get("url"))
    project_signals = devpost_project_signals(project)
    yield {
        "entityType": "project",
        "sourceNativeId": first_non_empty(project_url, project_name),
        "sourceUrl": project_url,
        "display": compact({
            "title": project_name,
            "tagline": project.get("tagline"),
            "hackathon": project.get("hackathon"),
            "homepage": project_url,
            "github": project.get("github_links") or project.get("github_repos"),
        }),
        "rawSummary": compact({
            "winner": project.get("winner"),
            "likes": project.get("likes"),
            "github": project.get("github_links") or project.get("github_repos"),
            "demo": project.get("demo_links"),
            "allLinks": project.get("all_links"),
            "video": project.get("video_url"),
            "tech": project.get("tech_tags") or project.get("tech_stack"),
            "prizes": project.get("prizes"),
            "suggestedSignals": project_signals,
        }),
        "raw": project,
    }

    for member in project.get("members", []) or []:
        if not isinstance(member, dict):
            continue
        member_name = first_non_empty(member.get("name"), member.get("member_name"))
        member_url = first_non_empty(member.get("devpost_profile"), member.get("member_devpost"))
        yield {
            "entityType": "person",
            "sourceNativeId": first_non_empty(member_url, member.get("github_url"), member_name),
            "sourceUrl": member_url,
            "displayName": member_name,
            "display": compact({
                "name": member_name,
                "homepage": member.get("website"),
                "github": member.get("github_url"),
                "linkedin": member.get("linkedin_url"),
            }),
            "rawSummary": compact({
                "devpost": member_url,
                "github": member.get("github_url"),
                "linkedin": member.get("linkedin_url"),
                "twitter": member.get("twitter_url"),
                "homepage": member.get("website"),
                "projectUrl": project_url,
                "projectName": project_name,
                "hackathon": project.get("hackathon"),
                "tech": project.get("tech_tags") or project.get("tech_stack"),
                "prizes": project.get("prizes"),
                "winner": project.get("winner"),
                "suggestedSignals": suggested_signals(*project_signals, "founder_or_builder_signal"),
            }),
            "raw": {
                "member": member,
                "project": compact({
                    "projectName": project_name,
                    "projectUrl": project_url,
                    "hackathon": project.get("hackathon"),
                    "tech": project.get("tech_tags") or project.get("tech_stack"),
                    "prizes": project.get("prizes"),
                    "winner": project.get("winner"),
                }),
            },
        }


def github_record(candidate: dict[str, Any]) -> dict[str, Any]:
    profile = candidate.get("profile") if isinstance(candidate.get("profile"), dict) else {}
    repos = candidate.get("repos") if isinstance(candidate.get("repos"), list) else []
    username = first_non_empty(candidate.get("username"), profile.get("login"))
    html_url = first_non_empty(profile.get("html_url"), f"https://github.com/{username}" if username else "")
    repo_names = unique_values([repo.get("full_name") for repo in repos if isinstance(repo, dict)])
    project_stars = sum(safe_int(repo.get("stars")) for repo in repos if isinstance(repo, dict))
    return {
        "entityType": "person",
        "sourceNativeId": username or html_url,
        "sourceUrl": html_url,
        "displayName": first_non_empty(profile.get("name"), username),
        "institution": profile.get("company"),
        "display": compact({
            "name": first_non_empty(profile.get("name"), username),
            "institution": profile.get("company"),
            "homepage": profile.get("blog"),
            "github": html_url,
        }),
        "rawSummary": compact({
            "email": first_non_empty(candidate.get("emails"), profile.get("email")),
            "github": html_url,
            "blog": profile.get("blog"),
            "bio": profile.get("bio"),
            "location": profile.get("location"),
            "followers": profile.get("followers"),
            "publicRepos": profile.get("public_repos"),
            "mergedPrs": candidate.get("merged_prs"),
            "totalCommits": candidate.get("total_commits"),
            "repoCount": len(repos),
            "projectStars": project_stars,
            "sourceRepos": repo_names,
            "score": candidate.get("score"),
            "suggestedSignals": github_suggested_signals(candidate),
        }),
        "raw": candidate,
    }


def generic_record(record: dict[str, Any]) -> dict[str, Any]:
    name = first_non_empty(
        record.get("name"),
        record.get("member_name"),
        record.get("username"),
        record.get("full_name"),
        record.get("project_name"),
    )
    institution = first_non_empty(record.get("institution"), record.get("company"), record.get("affiliation"))
    source_url = first_non_empty(
        record.get("github_url"),
        record.get("member_github"),
        record.get("html_url"),
        record.get("homepage"),
        record.get("url"),
        record.get("blog"),
        record.get("website"),
        record.get("member_website"),
        record.get("project_url"),
    )
    return {
        "entityType": first_non_empty(record.get("entityType"), record.get("entity_type"), record.get("type")) or "person",
        "sourceNativeId": first_non_empty(
            record.get("sourceNativeId"),
            record.get("source_native_id"),
            record.get("id"),
            record.get("username"),
            record.get("member_username"),
            record.get("member_devpost"),
            record.get("github_url"),
            record.get("member_github"),
            record.get("project_url"),
            record.get("email"),
            record.get("orcid"),
            record.get("doi"),
        ),
        "sourceUrl": source_url,
        "displayName": name,
        "institution": institution,
        "display": compact({
            "name": name,
            "title": first_non_empty(record.get("title"), record.get("project_name")),
            "institution": institution,
            "homepage": first_non_empty(record.get("homepage"), record.get("website"), record.get("member_website"), record.get("blog")),
            "github": first_non_empty(record.get("github_url"), record.get("member_github"), record.get("html_url")),
            "linkedin": first_non_empty(record.get("linkedin_url"), record.get("member_linkedin")),
        }),
        "rawSummary": compact({
            "email": first_non_empty(record.get("email"), record.get("emails"), record.get("member_email")),
            "orcid": first_non_empty(record.get("orcid"), record.get("member_orcid")),
            "github": first_non_empty(record.get("github_url"), record.get("member_github"), record.get("html_url")),
            "linkedin": first_non_empty(record.get("linkedin_url"), record.get("member_linkedin")),
            "devpost": first_non_empty(record.get("member_devpost"), record.get("project_url")),
            "doi": first_non_empty(record.get("doi"), record.get("paper_doi")),
            "venue": first_non_empty(record.get("venue"), record.get("journal")),
            "score": first_non_empty(record.get("score"), record.get("score_total")),
            "suggestedSignals": research_suggested_signals(record),
        }),
        "raw": record,
    }


def source_record_id(source: str, entity_type: str, native_id: str, raw: dict[str, Any]) -> str:
    stable = native_id or content_hash(raw).split(":", 1)[1][:16]
    return f"src_{safe_id(source)}_{safe_id(entity_type)}_{safe_id(stable)}"


def to_source_records(
    *,
    run_id: str,
    domain: str,
    source: str,
    records: Iterable[dict[str, Any]],
) -> list[dict[str, Any]]:
    now = utc_now()
    output: list[dict[str, Any]] = []
    for normalized in iter_domain_records(source, records):
        raw = normalized.get("raw") if isinstance(normalized.get("raw"), dict) else normalized
        entity_type = str(normalized.get("entityType") or "generic_record")
        native_id = first_non_empty(normalized.get("sourceNativeId"), normalized.get("sourceUrl"))
        record_id = source_record_id(source, entity_type, native_id, raw)
        output.append(compact({
            "sourceRecordId": record_id,
            "runId": run_id,
            "domain": domain,
            "source": source,
            "entityType": entity_type,
            "sourceNativeId": native_id or record_id,
            "sourceUrl": normalized.get("sourceUrl"),
            "displayName": normalized.get("displayName"),
            "institution": normalized.get("institution"),
            "display": normalized.get("display") or {},
            "rawSummary": normalized.get("rawSummary") or {},
            "raw": raw,
            "rawStoragePath": raw_storage_path(
                domain=domain,
                source=source,
                run_id=run_id,
                source_record_id=record_id,
            ),
            "contentHash": content_hash(raw),
            "schemaVersion": "sourcing_source_record.v1",
            "createdAt": now,
            "updatedAt": now,
        }))
    return output


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Upload Devpost, GitHub, researcher, or manual files to core-service sourcing API.")
    parser.add_argument("--input", required=True, type=Path, help="CSV, JSON, or JSONL file")
    parser.add_argument("--run-id", required=True)
    parser.add_argument("--domain", required=True, help="Example: researcher, developer, hackathon, manual")
    parser.add_argument("--source", required=True, help="Example: openalex, github, devpost, csv")
    parser.add_argument("--api-base-url", default="http://127.0.0.1:5100/api/sourcing")
    parser.add_argument("--output-root", default="data")
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE)
    parser.add_argument("--dry-run", action="store_true")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    raw_records = load_input_records(args.input)
    records = to_source_records(
        run_id=args.run_id,
        domain=args.domain,
        source=args.source,
        records=raw_records,
    )
    if not records:
        raise SystemExit(f"No records converted from {args.input}")

    if args.dry_run:
        result = dry_run_upload(
            data_root=args.output_root,
            run_id=args.run_id,
            domain=args.domain,
            records=records,
        )
        print(f"dry-run wrote {result.record_count} source records to {result.records_payload_path}")
        return 0

    result = upload_source_records(
        api_base_url=args.api_base_url,
        run_id=args.run_id,
        domain=args.domain,
        records=records,
        batch_size=args.batch_size,
    )
    print(f"uploaded {result.record_count} source records from {args.source} to {args.api_base_url}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
